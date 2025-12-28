import os
import pandas as pd
import openpyxl
from openpyxl.utils import range_boundaries
from pathlib import Path
from typing import Optional

def convert_xlsb_to_xlsx(file_path: Path) -> Path:
    """
    1. xlsb -> xlsx 변환
    """
    if str(file_path).endswith('.xlsx'):
        return file_path
        
    if not str(file_path).endswith('.xlsb'):
        return file_path

    # Read xlsb using pandas (requires pyxlsb)
    df = pd.read_excel(file_path, engine='pyxlsb')
    
    new_path = file_path.with_suffix('.xlsx')
    # Save as xlsx
    df.to_excel(new_path, index=False)
    
    # Optional: remove original file? Keeping it for now might be safer unless disk space is issue.
    # but requirement implies replacing/processing. 
    # Let's keep original for backup but return new path.
    return new_path

def preprocess_structure(file_path: Path, sheet_name: Optional[str] = None) -> openpyxl.Workbook:
    """
    2. 구조 전처리: B4부터 시작, 병합 셀 해제 및 값 채우기 (Forward Fill style)
    """
    wb = openpyxl.load_workbook(file_path)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # Start processing from row 4
    start_row = 4
    
    # 1. Unmerge all cells first to handle them individually
    # We need to collect merged ranges first because unmerging modifies the collection
    merged_ranges = list(ws.merged_cells.ranges)
    
    for merged_range in merged_ranges:
        # Get boundaries
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        
        # Only process if it affects our area of interest (Row >= 4)
        if max_row < start_row:
            continue
            
        # Get the value of the top-left cell
        top_left_value = ws.cell(row=min_row, column=min_col).value
        
        # Unmerge
        ws.unmerge_cells(str(merged_range))
        
        # Fill all cells in the range with the top-left value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if row >= start_row: # Valid data area
                    cell = ws.cell(row=row, column=col)
                    cell.value = top_left_value

    # 2. Also perform visual Forward Fill for empty cells in hierarchical columns if needed?
    # The requirement says "병합된 셀이 많아... 병합된 셀의 값을 채워야해".
    # Unmerging and filling top-left value handles the explicit merged cells.
    # If there are just empty cells that meant "ditto", that's harder to guess without explicit merge.
    # Assuming "merged cells" was the main issue.
    
    return wb

def consolidate_issue_column(wb: openpyxl.Workbook, sheet_name: Optional[str], issue_col_name: str = "Issue") -> Path:
    """
    3. Issue 컬럼 병합 처리
    - Issue 컬럼을 찾아서
    - 여러 행/열에 걸친 텍스트를 병합 (\n, space 구분)
    - 실제 셀도 병합
    """
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # Find Issue column index (assuming row 3 or 4 has headers, let's search in start_row - 1 or start_row)
    # Based on prev requirement "4B에서 테이블이 시작해" -> Row 4, Col 2 is start of data? Or Headers at row 3?
    # Let's assume headers are at row 3 (if data starts row 4).
    header_row = 3
    issue_col_idx = None
    
    for cell in ws[header_row]:
        if cell.value and str(cell.value).strip() == issue_col_name:
            issue_col_idx = cell.column
            break
    
    if not issue_col_idx:
        # Fallback: Search typical "Issue" column or return as is
        print("Issue column not found")
        return wb

    # Logic to identify "groups" of issues.
    # Requirement: "Issue 칸에 있는 여러 셀... 모델 칸과 같은데 실제론 여러 셀... 순서대로 병합"
    # This implies the "Model" column (or the primary key column) drives the grouping.
    # If Model is merged (and we unmerged/filled it in step 2), then we can group by Model.
    
    # We need to find the "Model" column or a grouping key.
    # Or, purely look at Issue column? 
    # "눈으로 보면 모델 칸과 같은데" means the visual height of the Issue cell matches the Model cell (which was merged).
    # Since we successfully propagated the Model value in Step 2, standard rows now have the same Model value.
    # So we can group consecutive rows that have the SAME Model value (and same Layer etc).
    
    # 1. Identify distinct groups based on key columns (Layer, Model etc to the left of Issue).
    # Assuming columns before Issue are keys.
    
    max_row = ws.max_row
    current_group_key = None
    current_group_rows = []
    
    # Collect data to process
    # We will modify sheet in-place, so maybe iterate first to calculate merges, then apply.
    
    # It's Complex to do in-place while iterating.
    # Strategy:
    # 1. Iterate rows 4 to Max.
    # 2. Build a key from columns 2 to Issue_Col-1.
    # 3. If key changes, finalize previous group.
    
    # Note: Column 2 (B) is start.
    cols_check = list(range(2, issue_col_idx)) # Columns B to right before Issue
    
    merged_regions = [] # (start_row, end_row, merged_text)
    
    for row in range(4, max_row + 2): # +2 to ensure flush of last group
        # Get Key
        if row <= max_row:
            key_values = [ws.cell(row=row, column=c).value for c in cols_check]
            key = tuple(key_values)
        else:
            key = "END_OF_DATA"

        if key != current_group_key:
            # Process previous group
            if current_group_rows:
                # Merge logic
                text_parts = []
                for r in current_group_rows:
                    cell_val = ws.cell(row=r, column=issue_col_idx).value
                    if cell_val:
                        text_parts.append(str(cell_val).strip())
                
                full_text = "\n".join(text_parts)
                
                # Record merge action
                if len(current_group_rows) > 0:
                    merged_regions.append({
                        "start_row": current_group_rows[0],
                        "end_row": current_group_rows[-1],
                        "text": full_text
                    })
            
            # Start new group
            current_group_key = key
            current_group_rows = [row]
        else:
            current_group_rows.append(row)
            
    # Apply merges
    api_alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='center')
    
    for region in merged_regions:
        s = region["start_row"]
        e = region["end_row"]
        txt = region["text"]
        
        # Set text to top-left
        main_cell = ws.cell(row=s, column=issue_col_idx)
        main_cell.value = txt
        main_cell.alignment = api_alignment
        
        # Merge if multiple rows
        if e > s:
            ws.merge_cells(start_row=s, start_column=issue_col_idx, end_row=e, end_column=issue_col_idx)
            
    return wb

def run_preprocessing_pipeline(file_path: Path, sheet_name: Optional[str] = None) -> Path:
    """
    파이프라인 실행
    """
    # 1. Convert
    xlsx_path = convert_xlsb_to_xlsx(file_path)
    
    try:
        # 2. Structure Preprocessing
        wb = preprocess_structure(xlsx_path, sheet_name=sheet_name)
        
        # 3. Issue Column Consolidation
        # Need to determine sheet name if not provided?
        if not sheet_name:
            sheet_name = wb.sheetnames[0]
            
        consolidate_issue_column(wb, sheet_name=sheet_name)
        
        # Save
        processed_path = xlsx_path.with_name(f"processed_{xlsx_path.name}")
        wb.save(processed_path)
        wb.close()
        
        return processed_path
        
    except Exception as e:
        print(f"Preprocessing failed: {e}")
        # Return original if fail? Or raise?
        raise e
