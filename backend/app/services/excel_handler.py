import polars as pl
from pathlib import Path
from typing import Optional, List, Dict, Any
import openpyxl



class ExcelHandler:
    """Excel 파일 처리를 위한 클래스"""
    
    @staticmethod
    def read_excel(
        file_path: str,
        sheet_name: str = "일보_DPU"
    ) -> pl.DataFrame:
        """
        Excel 파일을 읽어서 Polars DataFrame으로 반환
        
        Args:
            file_path: Excel 파일 경로
            sheet_name: 시트 이름
            
        Returns:
            Polars DataFrame
        """
        # Polars는 openpyxl 엔진을 사용하여 Excel 읽기
        df = pl.read_excel(
            file_path,
            sheet_name=sheet_name
        )
        return df
    
    @staticmethod
    def append_results_to_file(
        original_file_path: str,
        output_file_path: str,
        classifications: List[Dict[str, str]],
        sheet_name: str = "일보_DPU"
    ) -> str:
        """
        기존 엑셀 파일의 서식(병합 등)을 유지하면서 결과 컬럼 추가
        """
        # Load workbook
        wb = openpyxl.load_workbook(original_file_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Determine header row and keys
        # Assuming headers are at row 3 based on preprocessor logic, data starts row 4
        header_row = 3
        data_start_row = 4
        
        # New columns to add
        new_headers = ["불량명", "설비명", "조치내용"]
        
        # Find start column for new data (max column + 1)
        start_col = ws.max_column + 1
        
        # Write Headers
        for idx, header in enumerate(new_headers):
            cell = ws.cell(row=header_row, column=start_col + idx)
            cell.value = header
            # Simple styling
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
        # Write Data
        # classifications index 0 corresponds to data_start_row
        for i, result in enumerate(classifications):
            row_idx = data_start_row + i
            
            # Map values
            values = [
                result.get("불량명", ""),
                result.get("설비명", ""),
                result.get("조치내용", "")
            ]
            
            for col_offset, value in enumerate(values):
                cell = ws.cell(row=row_idx, column=start_col + col_offset)
                cell.value = value
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='center')

        # Save
        wb.save(output_file_path)
        wb.close()
        
        return output_file_path

    @staticmethod
    def write_excel(
        df: pl.DataFrame,
        file_path: str,
        sheet_name: str = "일보_DPU"
    ) -> str:
        """
        Polars DataFrame을 Excel 파일로 저장
        
        Args:
            df: 저장할 DataFrame
            file_path: 저장 경로
            sheet_name: 시트 이름
            
        Returns:
            저장된 파일 경로
        """
        df.write_excel(
            file_path,
            worksheet=sheet_name
        )
        return file_path
    
    @staticmethod
    def add_classification_columns(
        df: pl.DataFrame,
        classifications: List[Dict[str, str]]
    ) -> pl.DataFrame:
        """
        분류 결과를 DataFrame에 추가
        
        Args:
            df: 원본 DataFrame
            classifications: 분류 결과 리스트 (각 row에 대한 분류 결과)
            
        Returns:
            분류 컬럼이 추가된 DataFrame
        """
        # 분류 결과를 Series로 변환
        불량명_list = [c.get("불량명", "") for c in classifications]
        설비명_list = [c.get("설비명", "") for c in classifications]
        조치내용_list = [c.get("조치내용", "") for c in classifications]
        
        # 새 컬럼 추가
        df = df.with_columns([
            pl.Series("불량명", 불량명_list),
            pl.Series("설비명", 설비명_list),
            pl.Series("조치내용", 조치내용_list)
        ])
        
        return df
    
    @staticmethod
    def get_column_values(
        df: pl.DataFrame,
        column_name: str
    ) -> List[Optional[str]]:
        """
        특정 컬럼의 값들을 리스트로 반환
        
        Args:
            df: DataFrame
            column_name: 컬럼 이름
            
        Returns:
            컬럼 값 리스트
        """
        if column_name not in df.columns:
            raise ValueError(f"Column '{column_name}' not found in DataFrame")
        
        # null 값을 None으로 변환
        values = df[column_name].to_list()
        return values
    
    @staticmethod
    def is_empty_value(value: Any) -> bool:
        """값이 비어있는지 확인"""
        if value is None:
            return True
        if isinstance(value, str) and value.strip() == "":
            return True
        return False
